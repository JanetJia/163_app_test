from openpyxl import load_workbook


class ParseExcel:
    def __init__(self, excelpath, sheetname):
        self.wb = load_workbook(excelpath)
        self.sheet = self.wb.get_sheet_by_name(sheetname)
        self.max_column_num = self.sheet.max_column

    def getdatasfromsheet(self):
        datalist = []
        for line in self.sheet.rows[1:]:  # 手动跳过第一行
            temp_list = []
            for col in range(1, self.max_column_num-1):  # 手动跳过第一列
                temp_list.append(line[col].value)
            datalist.append(temp_list)
        return datalist[0:]
